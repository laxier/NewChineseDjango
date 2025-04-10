from django.views import generic
from .models import RelatedWord, Sentence
from chineseword.models import ChineseWord
from frontend.views import CurrentUserMixin


class ChineseWordDetailView(CurrentUserMixin, generic.DetailView):
    model = ChineseWord
    context_object_name = 'chinese_word'

    def get_template_names(self):
        is_mindmap = self.request.GET.get('is_mindmap', 'false') == 'true'
        if is_mindmap:
            return ['wordpages/chinese_word_detail_no_header.html']
        return ['wordpages/chinese_word_detail.html']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sentences'] = Sentence.objects.filter(chinese_word=self.object)
        context['related_words'] = RelatedWord.objects.filter(word=self.object)
        context['is_favorite'] = self.request.user in self.object.favorites.all()

        return context

import os
import datetime

from django.conf import settings
from django.http import HttpResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin

import openpyxl
from openpyxl.styles import PatternFill, Font

from users.models import WordPerformance
from chineseword.models import ChineseWord


class HSK4ProgressDownloadView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        input_path = os.path.join(settings.BASE_DIR, "progress_template_hsk4.xlsx")
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        # -----------------------------
        # СЧЁТЧИКИ ДЛЯ СЛОВ
        # -----------------------------
        total_words = 0
        none_count_words = 0
        low_count_words = 0
        medium_count_words = 0
        high_count_words = 0

        # -----------------------------
        # СЧЁТЧИКИ ДЛЯ ПРИМЕРОВ
        # -----------------------------
        total_examples = 0
        none_count_examples = 0
        low_count_examples = 0
        medium_count_examples = 0
        high_count_examples = 0

        # -----------------------------
        # 1) Собираем иероглифы
        # -----------------------------
        characters = set()

        # Для слов (столбец E=5 → index=4)
        for row in ws.iter_rows(min_row=2, max_col=15):
            word_cell = row[4]
            if word_cell.value:
                characters.add(word_cell.value)

        # Для примеров (столбец L=12 → index=11)
        for row in ws.iter_rows(min_row=2, max_col=15):
            example_cell = row[11]
            if example_cell.value:
                characters.add(example_cell.value)

        # -----------------------------
        # 2) Загружаем слова из БД
        # -----------------------------
        words_qs = ChineseWord.objects.filter(simplified__in=characters)
        word_dict = {w.simplified: w for w in words_qs}

        # -----------------------------
        # 3) Загружаем перформансы
        # -----------------------------
        user_id = request.user.id
        word_ids = [w.id for w in words_qs]
        perf_qs = WordPerformance.objects.filter(user_id=user_id, word_id__in=word_ids)
        perf_dict = {p.word_id: p for p in perf_qs}

        # -----------------------------
        # 4) Стили и функция раскраски
        # -----------------------------
        no_progress_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        no_progress_font = Font(color="000000")
        low_progress_font = Font(color="FF9900")
        good_progress_font = Font(color="008000")

        def color_cell_if_needed(cell, prog):
            if prog is None:
                cell.fill = no_progress_fill
            elif prog < 30:
                cell.font = no_progress_font
            elif 30 <= prog <= 80:
                cell.font = low_progress_font
            else:
                cell.font = good_progress_font

        # -----------------------------
        # 5) Проход по строкам и РЕАЛЬНЫЙ ПОДСЧЁТ СЛОВ
        # -----------------------------
        for row in ws.iter_rows(min_row=2, max_col=15):
            # СЛОВО в E (index=4)
            word_cell = row[4]
            if word_cell.value:
                word_obj = word_dict.get(word_cell.value)
                perf = perf_dict.get(word_obj.id) if word_obj else None

                # Считаем
                total_words += 1

                if perf and perf.accuracy_percentage_display is not None:
                    prog = perf.accuracy_percentage_display
                else:
                    prog = None

                # Логика распределения по категориям
                if prog is None:
                    none_count_words += 1
                elif prog < 30:
                    low_count_words += 1
                elif 30 <= prog <= 80:
                    medium_count_words += 1
                else:
                    high_count_words += 1

                # Запись в A,B,C
                wrong_count = perf.wrong if perf else 0
                right_count = perf.right if perf else 0
                row_num = word_cell.row
                a_cell = ws.cell(row=row_num, column=1)  # A
                b_cell = ws.cell(row=row_num, column=2)  # B
                c_cell = ws.cell(row=row_num, column=3)  # C

                a_cell.value = wrong_count
                b_cell.value = right_count
                c_cell.value = prog if prog is not None else ""

                # Раскраска
                color_cell_if_needed(word_cell, prog)

            # ПРИМЕР в L (index=11)
            example_cell = row[11]
            if example_cell.value:
                word_obj = word_dict.get(example_cell.value)
                perf = perf_dict.get(word_obj.id) if word_obj else None

                # Считаем
                total_examples += 1

                if perf and perf.accuracy_percentage_display is not None:
                    prog = perf.accuracy_percentage_display
                else:
                    prog = None

                # Логика распределения по категориям
                if prog is None:
                    none_count_examples += 1
                elif prog < 30:
                    low_count_examples += 1
                elif 30 <= prog <= 80:
                    medium_count_examples += 1
                else:
                    high_count_examples += 1

                # Запись в I,J,K (зависит от того, куда хотите писать)
                row_num = example_cell.row
                i_cell = ws.cell(row=row_num, column=9)   # I
                j_cell = ws.cell(row=row_num, column=10)  # J
                k_cell = ws.cell(row=row_num, column=11)  # K

                wrong_count = perf.wrong if perf else 0
                right_count = perf.right if perf else 0

                i_cell.value = wrong_count
                j_cell.value = right_count
                k_cell.value = prog if prog is not None else ""

                # Раскраска
                color_cell_if_needed(example_cell, prog)

        # -------------------------------------------------------
        # 6) Создаём лист Statistics и записываем результаты
        # -------------------------------------------------------
        current_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if "Statistics" in wb.sheetnames:
            stats_sheet = wb["Statistics"]
            for row in stats_sheet["A1:F50"]:
                for cell in row:
                    cell.value = None
        else:
            stats_sheet = wb.create_sheet("Statistics")

        stats_sheet["A1"] = "Дата и время"
        stats_sheet["B1"] = current_dt
        stats_sheet.column_dimensions['A'].width = 25
        stats_sheet.column_dimensions['B'].width = 15

        # Слова
        stats_sheet["A2"] = "Слова"
        stats_sheet["A3"] = "Всего слов"
        stats_sheet["B3"] = total_words

        stats_sheet["A4"] = "Без прогресса (None)"
        stats_sheet["B4"] = none_count_words

        stats_sheet["A5"] = "Прогресс 0-30"
        stats_sheet["B5"] = low_count_words

        stats_sheet["A6"] = "Прогресс 30-80"
        stats_sheet["B6"] = medium_count_words

        stats_sheet["A7"] = "Прогресс >80"
        stats_sheet["B7"] = high_count_words

        # Примеры
        stats_sheet["A10"] = "Примеры"
        stats_sheet["A11"] = "Всего примеров"
        stats_sheet["B11"] = total_examples

        stats_sheet["A12"] = "Без прогресса (None)"
        stats_sheet["B12"] = none_count_examples

        stats_sheet["A13"] = "Прогресс 0-30"
        stats_sheet["B13"] = low_count_examples

        stats_sheet["A14"] = "Прогресс 30-80"
        stats_sheet["B14"] = medium_count_examples

        stats_sheet["A15"] = "Прогресс >80"
        stats_sheet["B15"] = high_count_examples

        # ---------------------------------------------------
        # 7) Сохраняем и отдаём файл
        # ---------------------------------------------------
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"hsk4_progress_{timestamp}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
